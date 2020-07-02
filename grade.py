import pandas as pd
import numpy as np
from openpyxl import load_workbook

def load_grade(path):
	dt = pd.read_csv(path, sep=';', dtype={'matricule': object})
	return dt.set_index('matricule')

def capwords(S):
	return ' '.join([w.capitalize() for w in S.split(' ')])

def save_grade(df, path):
	 df.to_csv(path, sep=';', index_label='matricule')

def fill_excel(df, excel, firstRow=4, matriculeColumn='B', gradeColumn='D', statusColumn='E', dfColumn='grade'):
	wb = load_workbook(filename=excel)
	sheets = list(wb.sheetnames)
	sheets.remove('NE_PAS_EFFACER')
	ws = wb[sheets[0]]
	lastRow = int(''.join(list(filter(lambda l: l in '0123456789', ws.calculate_dimension().split(':')[1]))))
	index = {}
	for i in range(firstRow, lastRow+1):
		index[ws[matriculeColumn+str(i)].value] = gradeColumn+str(i)
	
	def fun(row):
		matricule = int(row.name)
		if matricule in index:
			if ws[index[matricule]].value is None:
				ws[index[matricule]] = row[dfColumn]
			else:
				print(f'{matricule} already get grade')
			del index[matricule]
		else:
			print(f'{matricule} not in file')

	df.apply(fun, axis=1)

	for matricule in index:
		print(f'{matricule} don\'t have grade => absent')
		ws[index[matricule]] = 0.0
		status = statusColumn + ''.join(list(filter(lambda l: l in '0123456789', index[matricule])))
		ws[status] = 'a'

	wb.save(filename=excel)

def merge_grades(**kwargs):
	out = pd.DataFrame()
	for col, df in kwargs.items():
		out = out.join(df.rename(columns={'grade': col}), how='outer', rsuffix='_'+col)
	
	return out

